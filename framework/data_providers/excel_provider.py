"""Excel data provider for .xlsx and .xls files."""

import openpyxl
import xlrd
from pathlib import Path
from typing import Any, Dict, List, Optional, Union
from framework.data_providers.base_provider import BaseDataProvider


class ExcelProvider(BaseDataProvider):
    """Provider for Excel data sources (.xlsx, .xls)."""
    
    def __init__(self, source: str):
        """
        Initialize Excel provider.
        
        Args:
            source: Path to Excel file
        """
        super().__init__(source)
        self.file_path = Path(source)
        self.workbook = None
        self.file_type = self.file_path.suffix.lower()
        
        if not self.file_path.exists():
            raise FileNotFoundError(f"Excel file not found: {source}")
        
        self._load_workbook()
    
    def _load_workbook(self) -> None:
        """Load Excel workbook based on file type."""
        if self.file_type == ".xlsx":
            self.workbook = openpyxl.load_workbook(self.file_path, data_only=True)
        elif self.file_type == ".xls":
            self.workbook = xlrd.open_workbook(self.file_path)
        else:
            raise ValueError(f"Unsupported file type: {self.file_type}")
    
    def get_data(
        self,
        sheet: Optional[Union[str, int]] = 0,
        row: Optional[int] = None,
        column: Optional[Union[str, int]] = None,
        cell: Optional[str] = None,
        row_as_dict: bool = True
    ) -> Any:
        """
        Get data from Excel file.
        
        Args:
            sheet: Sheet name or index (default: 0)
            row: Row number (1-based)
            column: Column name or index
            cell: Cell address (e.g., 'A1')
            row_as_dict: Return row as dictionary with headers
            
        Returns:
            Cell value, row data, or column data
        """
        if self.file_type == ".xlsx":
            return self._get_data_xlsx(sheet, row, column, cell, row_as_dict)
        else:
            return self._get_data_xls(sheet, row, column, cell, row_as_dict)
    
    def _get_data_xlsx(
        self,
        sheet: Union[str, int],
        row: Optional[int],
        column: Optional[Union[str, int]],
        cell: Optional[str],
        row_as_dict: bool
    ) -> Any:
        """Get data from .xlsx file."""
        # Get worksheet
        if isinstance(sheet, int):
            ws = self.workbook.worksheets[sheet]
        else:
            ws = self.workbook[sheet]
        
        # Get specific cell
        if cell:
            return ws[cell].value
        
        # Get entire row
        if row and not column:
            row_data = []
            for cell in ws[row]:
                row_data.append(cell.value)
            
            if row_as_dict and row > 1:
                headers = [cell.value for cell in ws[1]]
                return dict(zip(headers, row_data))
            return row_data
        
        # Get entire column
        if column and not row:
            col_data = []
            if isinstance(column, str):
                for cell in ws[column]:
                    col_data.append(cell.value)
            else:
                for row in ws.iter_rows(min_col=column, max_col=column):
                    col_data.append(row[0].value)
            return col_data
        
        # Get specific cell by row and column
        if row and column:
            if isinstance(column, str):
                return ws[f"{column}{row}"].value
            else:
                return ws.cell(row=row, column=column).value
        
        return None
    
    def _get_data_xls(
        self,
        sheet: Union[str, int],
        row: Optional[int],
        column: Optional[Union[str, int]],
        cell: Optional[str],
        row_as_dict: bool
    ) -> Any:
        """Get data from .xls file."""
        # Get worksheet
        if isinstance(sheet, int):
            ws = self.workbook.sheet_by_index(sheet)
        else:
            ws = self.workbook.sheet_by_name(sheet)
        
        # Get specific cell
        if cell:
            # Parse cell address (e.g., 'A1')
            col = ord(cell[0].upper()) - ord('A')
            row_num = int(cell[1:]) - 1
            return ws.cell_value(row_num, col)
        
        # Get entire row
        if row and not column:
            row_idx = row - 1
            row_data = ws.row_values(row_idx)
            
            if row_as_dict and row > 1:
                headers = ws.row_values(0)
                return dict(zip(headers, row_data))
            return row_data
        
        # Get entire column
        if column and not row:
            if isinstance(column, str):
                col_idx = ord(column.upper()) - ord('A')
            else:
                col_idx = column - 1
            return ws.col_values(col_idx)
        
        # Get specific cell by row and column
        if row and column:
            row_idx = row - 1
            if isinstance(column, str):
                col_idx = ord(column.upper()) - ord('A')
            else:
                col_idx = column - 1
            return ws.cell_value(row_idx, col_idx)
        
        return None
    
    def get_all_data(
        self,
        sheet: Optional[Union[str, int]] = 0,
        has_header: bool = True,
        start_row: int = 2
    ) -> List[Dict[str, Any]]:
        """
        Get all data from a sheet.
        
        Args:
            sheet: Sheet name or index
            has_header: Whether first row contains headers
            start_row: Starting row for data (1-based)
            
        Returns:
            List of dictionaries containing row data
        """
        if self.file_type == ".xlsx":
            return self._get_all_data_xlsx(sheet, has_header, start_row)
        else:
            return self._get_all_data_xls(sheet, has_header, start_row)
    
    def _get_all_data_xlsx(
        self,
        sheet: Union[str, int],
        has_header: bool,
        start_row: int
    ) -> List[Dict[str, Any]]:
        """Get all data from .xlsx file."""
        if isinstance(sheet, int):
            ws = self.workbook.worksheets[sheet]
        else:
            ws = self.workbook[sheet]
        
        data = []
        headers = []
        
        if has_header:
            headers = [cell.value for cell in ws[1]]
        
        for row in ws.iter_rows(min_row=start_row, values_only=True):
            if has_header:
                data.append(dict(zip(headers, row)))
            else:
                data.append(list(row))
        
        return data
    
    def _get_all_data_xls(
        self,
        sheet: Union[str, int],
        has_header: bool,
        start_row: int
    ) -> List[Dict[str, Any]]:
        """Get all data from .xls file."""
        if isinstance(sheet, int):
            ws = self.workbook.sheet_by_index(sheet)
        else:
            ws = self.workbook.sheet_by_name(sheet)
        
        data = []
        headers = []
        
        if has_header:
            headers = ws.row_values(0)
        
        for row_idx in range(start_row - 1, ws.nrows):
            row_data = ws.row_values(row_idx)
            if has_header:
                data.append(dict(zip(headers, row_data)))
            else:
                data.append(row_data)
        
        return data
    
    def get_sheet_names(self) -> List[str]:
        """
        Get all sheet names in the workbook.
        
        Returns:
            List of sheet names
        """
        if self.file_type == ".xlsx":
            return self.workbook.sheetnames
        else:
            return self.workbook.sheet_names()
    
    def get_row_count(self, sheet: Union[str, int] = 0) -> int:
        """
        Get row count for a sheet.
        
        Args:
            sheet: Sheet name or index
            
        Returns:
            Number of rows
        """
        if self.file_type == ".xlsx":
            if isinstance(sheet, int):
                ws = self.workbook.worksheets[sheet]
            else:
                ws = self.workbook[sheet]
            return ws.max_row
        else:
            if isinstance(sheet, int):
                ws = self.workbook.sheet_by_index(sheet)
            else:
                ws = self.workbook.sheet_by_name(sheet)
            return ws.nrows
    
    def close(self) -> None:
        """Close the workbook."""
        if self.workbook and self.file_type == ".xlsx":
            self.workbook.close()